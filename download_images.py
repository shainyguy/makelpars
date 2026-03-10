import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from PIL import Image as PILImage
import time
import os

# Файлы
OLD_FILE = "old.xlsx"
NEW_FILE = "new.xlsx"
OUTPUT_FILE = "new_with_images.xlsx"

# Настройки
HEADERS = {"User-Agent":"Mozilla/5.0"}
START_ROW = 2
SLEEP_TIME = 1

# Размер ячейки (приблизительно)
CELL_WIDTH = 100
CELL_HEIGHT = 100

# Загрузка Excel
df_new = pd.read_excel(NEW_FILE)
wb_new = load_workbook(NEW_FILE)
ws_new = wb_new.active

# Загрузка старого файла
wb_old = load_workbook(OLD_FILE)
ws_old = wb_old.active

# Словарь существующих картинок
existing_images = {}
for shp in ws_old._images:
    try:
        row = shp.anchor._from.row + 1
        article = ws_old.cell(row=row, column=2).value
        if article:
            existing_images[str(article).strip()] = shp
    except:
        continue

print(f"Найдено {len(existing_images)} старых картинок")

# Функция поиска картинки на сайте
def get_image_from_site(search_url, img_selector="img"):
    try:
        r = requests.get(search_url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "html.parser")
        img = soup.select_one(img_selector)
        if img and img.get("src") and img["src"].startswith("http"):
            return img["src"]
    except:
        return None
    return None

# Функция масштабирования под ячейку
def scale_image_to_cell(pil_img):
    img_w, img_h = pil_img.size
    scale_w = CELL_WIDTH / img_w
    scale_h = CELL_HEIGHT / img_h
    scale = min(scale_w, scale_h, 1)  # не увеличиваем больше 100%
    new_w = int(img_w * scale)
    new_h = int(img_h * scale)
    return pil_img.resize((new_w, new_h), PILImage.Resampling.LANCZOS)

# Основной цикл
for i, article in enumerate(df_new.iloc[:,1], start=START_ROW):
    if pd.isna(article):
        continue
    article_str = str(article).strip()

    # 1. Старая картинка
    if article_str in existing_images:
        shp_old = existing_images[article_str]
        try:
            temp_file = f"temp_{article_str}.png"
            with open(temp_file, "wb") as f:
                f.write(shp_old._data())
            pil_img = PILImage.open(temp_file)
            pil_img = scale_image_to_cell(pil_img)
            buffer = BytesIO()
            pil_img.save(buffer, format="JPEG")
            buffer.seek(0)
            picture = XLImage(buffer)
            ws_new.add_image(picture, f"A{i}")
            os.remove(temp_file)
            print(f"[OLD] Картинка перенесена: {article_str}")
            continue
        except Exception as e:
            print(f"[ERROR] Не удалось вставить старую картинку: {article_str}, причина: {e}")

    # 2. Поиск на сайтах
    img_url = None
    sites = [
        f"https://www.vseinstrumenti.ru/search/?q={article_str}+makel",
        f"https://www.petrovich.ru/search/?q={article_str}+makel",
        f"https://rs24.ru/search/?q={article_str}+makel"
    ]
    for site in sites:
        img_url = get_image_from_site(site)
        if img_url:
            break

    # 3. Вставка картинки
    if img_url:
        try:
            img_data = requests.get(img_url, headers=HEADERS, timeout=10).content
            pil_img = PILImage.open(BytesIO(img_data)).convert("RGB")
            pil_img = scale_image_to_cell(pil_img)
            buffer = BytesIO()
            pil_img.save(buffer, format="JPEG")
            buffer.seek(0)
            picture = XLImage(buffer)
            ws_new.add_image(picture, f"A{i}")
            print(f"[WEB] Картинка загружена: {article_str}")
        except Exception as e:
            print(f"[ERROR] Не удалось вставить: {article_str}, причина: {e}")
    else:
        print(f"[NOT FOUND] Картинка не найдена: {article_str}")

    time.sleep(SLEEP_TIME)

# Сохраняем Excel
wb_new.save(OUTPUT_FILE)
print("Готово! Все картинки обработаны и сохранены в", OUTPUT_FILE)
